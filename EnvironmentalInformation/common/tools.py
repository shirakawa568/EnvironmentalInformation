import os

import pandas

from openpyxl import load_workbook


def get_root_path(project_name):
    cur_path = os.path.abspath(os.path.dirname(__file__))
    root_path = cur_path[:cur_path.find(project_name + "\\") + len(project_name + "\\")]
    return root_path


def check_sheet(path, file_name, sheet_name):
    excel = pandas.ExcelFile(path + file_name)
    if sheet_name in excel.sheet_names:
        excel.close()
        return True
    else:
        excel.close()
        return False


def add_sheet(path, file_name, sheet_name, dataframe):
    """
    对已存在sheet的文件保留原sheet，追加新sheet
    :param path:
    :param file_name:
    :param sheet_name:
    :param dataframe:
    :return:
    """
    # 读取原sheet
    book = load_workbook(path + file_name)
    # 创建writer
    writer = pandas.ExcelWriter(path + file_name, engine='openpyxl')
    # 清空原sheet数据
    if check_sheet(path, file_name, sheet_name):
        book.remove(book.worksheets[book.sheetnames.index(sheet_name)])
    # 回写入Excel
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # 写入新sheet
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()


if __name__ == '__main__':
    df = pandas.DataFrame(columns=["aaa"])
    print(add_sheet(r"E:\_Project\Project_2020\EnvironmentalInformation\\", "Enterprises.xlsx", "企业详细信息", df))
